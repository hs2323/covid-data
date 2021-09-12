import openpyxl
import sys

from .country import Country

def open_excel(path):
  """Opens an excel file

  Args:
      path (str): The file location of the spreadsheet

  Returns:
      list: a list of Country objects
  """

  print("Opening data excel file \n")

  try:
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    data = []

    for row in worksheet.iter_rows(min_row=2):
      try:
        date = row[0].value.strftime("%Y-%m-%d")
        iso = row[1].value
        data.append(Country(date=date, iso=iso))
      except AttributeError:
        print ("Issue parsing row with ISO: " + row[1].value + " and Date:" + row[0].value + "\n")
 
    return data
  except openpyxl.utils.exceptions.InvalidFileException:
    print ("The data file does not exist or is in an invalid format \n")
    sys.exit("Exiting application")

def write_excel(path, list):
  """Writes an excel file

  Args:
      path (str): The file location of the spreadsheet
      list: array of Country objects with API data

  Returns:
      none
  """
  workbook = openpyxl.Workbook()
  sheet = workbook.active

  sheet.cell(row=1, column=1).value = 'date'
  sheet.cell(row=1, column=2).value = 'iso'
  sheet.cell(row=1, column=3).value = 'num_confirmed'
  sheet.cell(row=1, column=4).value = 'num_deaths'
  sheet.cell(row=1, column=5).value = 'num_recovered'

  for index, country in enumerate(list):
    row = index + 2
    sheet.cell(row=row, column=1).value = country.date
    sheet.cell(row=row, column=2).value = country.iso
    sheet.cell(row=row, column=3).value = country.confirmed
    sheet.cell(row=row, column=4).value = country.deaths
    sheet.cell(row=row, column=5).value = country.recovered

  try:
    workbook.save(path)
  except FileNotFoundError:
    print('Could not save data file in the location provided \n')
    sys.exit("Exiting application")

  print('Saved COVID API data at' + path)
